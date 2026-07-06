from pathlib import Path
import csv

from openpyxl import load_workbook

from .models import (
    CourseImport,
    ImportPackage,
    RoomImport,
    SchoolConfigurationImport,
    SubjectImport,
    TeacherImport,
)


class OfficialFileParser:
    """Parser de archivos oficiales/exportados.

    Soporta CSV y Excel con hojas:
    - Profesores
    - Cursos
    - Materias
    - Aulas
    - Centro

    Es una capa flexible para archivos oficiales o exportaciones autorizadas.
    """

    def parse_file(self, path: str | Path) -> ImportPackage:
        path = Path(path)
        suffix = path.suffix.lower()

        if suffix == ".csv":
            return self._parse_csv(path)

        if suffix in {".xlsx", ".xlsm"}:
            return self._parse_excel(path)

        raise ValueError(f"Formato no soportado: {suffix}")

    def _parse_csv(self, path: Path) -> ImportPackage:
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            rows = list(csv.DictReader(fh))

        teachers = []
        courses = []
        subjects = []
        rooms = []

        for row in rows:
            kind = self._value(row, "tipo").lower()
            if kind == "profesor":
                teachers.append(self._teacher(row))
            elif kind == "curso":
                courses.append(self._course(row))
            elif kind == "materia":
                subjects.append(self._subject(row))
            elif kind == "aula":
                rooms.append(self._room(row))

        return ImportPackage(
            teachers=tuple(teachers),
            courses=tuple(courses),
            subjects=tuple(subjects),
            rooms=tuple(rooms),
            source=str(path),
        )

    def _parse_excel(self, path: Path) -> ImportPackage:
        wb = load_workbook(path, data_only=True)

        teachers = self._parse_sheet(wb, "Profesores", self._teacher)
        courses = self._parse_sheet(wb, "Cursos", self._course)
        subjects = self._parse_sheet(wb, "Materias", self._subject)
        rooms = self._parse_sheet(wb, "Aulas", self._room)
        configuration = self._parse_configuration(wb)

        return ImportPackage(
            teachers=tuple(teachers),
            courses=tuple(courses),
            subjects=tuple(subjects),
            rooms=tuple(rooms),
            configuration=configuration,
            source=str(path),
        )

    def _parse_sheet(self, wb, sheet_name: str, mapper):
        if sheet_name not in wb.sheetnames:
            return []

        ws = wb[sheet_name]
        headers = [
            str(cell.value).strip().lower() if cell.value is not None else ""
            for cell in ws[1]
        ]

        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(value is None or str(value).strip() == "" for value in row):
                continue
            data = dict(zip(headers, row))
            items.append(mapper(data))

        return items

    def _parse_configuration(self, wb):
        if "Centro" not in wb.sheetnames:
            return None

        ws = wb["Centro"]
        data = {}
        for key, value in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if key is None:
                continue
            data[str(key).strip().lower()] = value

        return SchoolConfigurationImport(
            center_name=str(data.get("nombre", "") or ""),
            center_code=str(data.get("codigo", "") or ""),
            locality=str(data.get("localidad", "") or ""),
            province=str(data.get("provincia", "") or ""),
            sessions_per_day=int(data.get("sesiones_dia", 6) or 6),
            session_duration_minutes=int(data.get("duracion_sesion", 45) or 45),
            break_after_period=int(data.get("recreo_tras", 3) or 3),
            break_duration_minutes=int(data.get("duracion_recreo", 30) or 30),
            start_time=str(data.get("hora_inicio", "09:00") or "09:00"),
        )

    def _teacher(self, row) -> TeacherImport:
        return TeacherImport(
            code=self._value(row, "codigo"),
            name=self._value(row, "nombre"),
            surname=self._value(row, "apellidos"),
            speciality=self._value(row, "especialidad"),
            weekly_hours=int(self._value(row, "horas_semanales", 25) or 25),
            role=self._value(row, "cargo"),
        )

    def _course(self, row) -> CourseImport:
        return CourseImport(
            code=self._value(row, "codigo"),
            stage=self._value(row, "etapa", "Primaria"),
            level=int(self._value(row, "nivel", 1) or 1),
            group_name=self._value(row, "grupo", "A"),
            students=int(self._value(row, "alumnado", 25) or 25),
        )

    def _subject(self, row) -> SubjectImport:
        return SubjectImport(
            code=self._value(row, "codigo"),
            name=self._value(row, "nombre"),
            weekly_sessions=int(self._value(row, "sesiones_semanales", 1) or 1),
            speciality=self._value(row, "especialidad"),
            room_type=self._value(row, "tipo_aula", "Ordinaria"),
        )

    def _room(self, row) -> RoomImport:
        return RoomImport(
            code=self._value(row, "codigo"),
            name=self._value(row, "nombre"),
            room_type=self._value(row, "tipo", "Ordinaria"),
            capacity=int(self._value(row, "capacidad", 25) or 25),
        )

    def _value(self, row, key: str, default=""):
        value = row.get(key)
        if value is None:
            return default
        return str(value).strip()
