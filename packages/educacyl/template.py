from pathlib import Path
from openpyxl import Workbook


class OfficialImportTemplate:
    def create(self, path: str | Path) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        ws = wb.active
        ws.title = "Profesores"
        ws.append(["codigo", "nombre", "apellidos", "especialidad", "horas_semanales", "cargo"])
        ws.append(["P01", "Ana", "García", "Primaria", 25, "Tutoría"])

        ws = wb.create_sheet("Cursos")
        ws.append(["codigo", "etapa", "nivel", "grupo", "alumnado"])
        ws.append(["1A", "Primaria", 1, "A", 22])

        ws = wb.create_sheet("Materias")
        ws.append(["codigo", "nombre", "sesiones_semanales", "especialidad", "tipo_aula"])
        ws.append(["LEN", "Lengua", 5, "Primaria", "Ordinaria"])

        ws = wb.create_sheet("Aulas")
        ws.append(["codigo", "nombre", "tipo", "capacidad"])
        ws.append(["A1", "Aula 1", "Ordinaria", 25])

        ws = wb.create_sheet("Centro")
        ws.append(["campo", "valor"])
        ws.append(["nombre", "CEIP Tierra de Pinares"])
        ws.append(["codigo", "47001559"])
        ws.append(["localidad", "Mojados"])
        ws.append(["provincia", "Valladolid"])
        ws.append(["sesiones_dia", 6])
        ws.append(["duracion_sesion", 45])
        ws.append(["recreo_tras", 3])
        ws.append(["duracion_recreo", 30])
        ws.append(["hora_inicio", "09:00"])

        wb.save(path)
        return path
