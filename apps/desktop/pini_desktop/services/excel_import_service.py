from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook

from pini_desktop.services.availability_service import AvailabilityService, AvailabilityStatus
from pini_desktop.services.course_service import Course, CourseService
from pini_desktop.services.course_subject_service import CourseSubject, CourseSubjectService
from pini_desktop.services.dynamic_rule_service import DynamicRule, DynamicRuleService
from pini_desktop.services.room_service import Room, RoomService
from pini_desktop.services.subject_service import Subject, SubjectService
from pini_desktop.services.teacher_service import Teacher, TeacherService


@dataclass(frozen=True)
class ImportResult:
    created_teachers: int = 0
    created_courses: int = 0
    created_subjects: int = 0
    created_rooms: int = 0
    created_course_subjects: int = 0
    updated_availability: int = 0
    created_dynamic_rules: int = 0
    errors: tuple[str, ...] = ()

    @property
    def ok(self) -> bool:
        return not self.errors


class ExcelImportService:
    TEACHERS_SHEET = "Profesores"
    COURSES_SHEET = "Cursos"
    SUBJECTS_SHEET = "Materias"
    ROOMS_SHEET = "Aulas"
    COURSE_SUBJECTS_SHEET = "MateriasCurso"
    AVAILABILITY_SHEET = "Disponibilidad"
    DYNAMIC_RULES_SHEET = "Reglas"

    def __init__(self, database_path=None):
        self.teacher_service = TeacherService(database_path) if database_path else TeacherService()
        self.course_service = CourseService(database_path) if database_path else CourseService()
        self.subject_service = SubjectService(database_path) if database_path else SubjectService()
        self.room_service = RoomService(database_path) if database_path else RoomService()
        self.course_subject_service = CourseSubjectService(database_path) if database_path else CourseSubjectService()
        self.availability_service = AvailabilityService(database_path) if database_path else AvailabilityService()
        self.dynamic_rule_service = DynamicRuleService(database_path) if database_path else DynamicRuleService()

    def create_template(self, path: str | Path) -> Path:
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        ws = wb.active
        ws.title = self.TEACHERS_SHEET
        ws.append(["codigo", "nombre", "apellidos", "especialidad", "horas_semanales", "max_sesiones_dia"])
        ws.append(["P01", "Ana", "García", "Primaria", 25, 5])

        ws = wb.create_sheet(self.COURSES_SHEET)
        ws.append(["codigo", "etapa", "nivel", "grupo", "alumnado"])
        ws.append(["1A", "Primaria", 1, "A", 22])

        ws = wb.create_sheet(self.SUBJECTS_SHEET)
        ws.append(["codigo", "nombre", "sesiones_semanales", "especialidad", "tipo_aula", "max_consecutivas", "doble_sesion"])
        ws.append(["LEN", "Lengua", 5, "Primaria", "Ordinaria", 1, "No"])
        ws.append(["ING", "Inglés", 3, "Inglés", "Ordinaria", 2, "Sí"])
        ws.append(["CON", "Contextos", 3, "Primaria", "Ordinaria", 1, "No"])

        ws = wb.create_sheet(self.ROOMS_SHEET)
        ws.append(["codigo", "nombre", "tipo", "capacidad", "edificio", "recursos", "disponible"])
        ws.append(["A1", "Aula 1", "Ordinaria", 25, "Principal", "", "Sí"])

        ws = wb.create_sheet(self.COURSE_SUBJECTS_SHEET)
        ws.append(["curso_codigo", "materia_codigo", "sesiones_semanales", "profesor_codigo", "tipo_aula", "notas"])
        ws.append(["1A", "LEN", 5, "P01", "Ordinaria", ""])
        ws.append(["1A", "ING", 3, "", "Ordinaria", ""])

        ws = wb.create_sheet(self.AVAILABILITY_SHEET)
        ws.append(["profesor_codigo", "dia", "periodo", "estado"])
        ws.append(["P01", 1, 1, "AVAILABLE"])
        ws.append(["P01", 1, 6, "FORBIDDEN"])

        ws = wb.create_sheet(self.DYNAMIC_RULES_SHEET)
        ws.append(["codigo", "nombre", "ambito", "tipo", "prioridad", "afecta_a", "dia", "periodo", "valor", "activa", "notas"])
        ws.append(["R001", "Máximo general 1 consecutiva", "Materia", "Máximo consecutivas", "Obligatoria", "Todas", "", "", "1", "Sí", "Regla general del centro"])
        ws.append(["R002", "Inglés 4º-6º permite 2 consecutivas", "Materia", "Máximo consecutivas", "Obligatoria", "Inglés 4º-6º", "", "", "2", "Sí", "Excepción"])
        ws.append(["R003", "Contextos después del recreo", "Materia", "Después del recreo", "Obligatoria", "Contextos", "", "", "Sí", "Sí", "Regla del centro"])

        wb.save(path)
        return path

    def import_workbook(self, path: str | Path) -> ImportResult:
        wb = load_workbook(path)
        errors: list[str] = []
        counters = {
            "teachers": 0,
            "courses": 0,
            "subjects": 0,
            "rooms": 0,
            "course_subjects": 0,
            "availability": 0,
            "dynamic_rules": 0,
        }

        steps = [
            (self.TEACHERS_SHEET, self._import_teachers, "teachers"),
            (self.COURSES_SHEET, self._import_courses, "courses"),
            (self.SUBJECTS_SHEET, self._import_subjects, "subjects"),
            (self.ROOMS_SHEET, self._import_rooms, "rooms"),
            (self.COURSE_SUBJECTS_SHEET, self._import_course_subjects, "course_subjects"),
            (self.AVAILABILITY_SHEET, self._import_availability, "availability"),
            (self.DYNAMIC_RULES_SHEET, self._import_dynamic_rules, "dynamic_rules"),
        ]

        for sheet_name, importer, key in steps:
            if sheet_name in wb.sheetnames:
                created, sheet_errors = importer(wb[sheet_name])
                counters[key] += created
                errors.extend(sheet_errors)

        return ImportResult(
            created_teachers=counters["teachers"],
            created_courses=counters["courses"],
            created_subjects=counters["subjects"],
            created_rooms=counters["rooms"],
            created_course_subjects=counters["course_subjects"],
            updated_availability=counters["availability"],
            created_dynamic_rules=counters["dynamic_rules"],
            errors=tuple(errors),
        )

    def _rows(self, worksheet):
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in worksheet[1]]
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(value is None or str(value).strip() == "" for value in row):
                continue
            yield row_index, dict(zip(headers, row))

    def _teacher_map(self):
        return {teacher.code: teacher for teacher in self.teacher_service.list_teachers()}

    def _course_map(self):
        return {course.code: course for course in self.course_service.list_courses()}

    def _subject_map(self):
        return {subject.code: subject for subject in self.subject_service.list_subjects()}

    def _import_teachers(self, worksheet):
        created, errors = 0, []
        existing = {teacher.code for teacher in self.teacher_service.list_teachers()}
        for row_index, row in self._rows(worksheet):
            try:
                code = str(row.get("codigo") or "").strip()
                if not code or code in existing:
                    continue
                self.teacher_service.create_teacher(
                    Teacher(None, code, str(row.get("nombre") or "").strip(), str(row.get("apellidos") or "").strip(), str(row.get("especialidad") or "").strip(), int(row.get("horas_semanales") or 25), int(row.get("max_sesiones_dia") or 5))
                )
                existing.add(code)
                created += 1
            except Exception as exc:
                errors.append(f"Profesores fila {row_index}: {exc}")
        return created, errors

    def _import_courses(self, worksheet):
        created, errors = 0, []
        existing = {course.code for course in self.course_service.list_courses()}
        for row_index, row in self._rows(worksheet):
            try:
                code = str(row.get("codigo") or "").strip()
                if not code or code in existing:
                    continue
                self.course_service.create_course(Course(None, code, str(row.get("etapa") or "Primaria").strip(), int(row.get("nivel") or 1), str(row.get("grupo") or "A").strip(), int(row.get("alumnado") or 25)))
                existing.add(code)
                created += 1
            except Exception as exc:
                errors.append(f"Cursos fila {row_index}: {exc}")
        return created, errors

    def _import_subjects(self, worksheet):
        created, errors = 0, []
        existing = {subject.code for subject in self.subject_service.list_subjects()}
        for row_index, row in self._rows(worksheet):
            try:
                code = str(row.get("codigo") or "").strip().upper()
                if not code or code in existing:
                    continue
                self.subject_service.create_subject(
                    Subject(None, code, str(row.get("nombre") or "").strip(), int(row.get("sesiones_semanales") or 1), str(row.get("especialidad") or "").strip(), str(row.get("tipo_aula") or "").strip(), int(row.get("max_consecutivas") or 1), self._bool(row.get("doble_sesion")))
                )
                existing.add(code)
                created += 1
            except Exception as exc:
                errors.append(f"Materias fila {row_index}: {exc}")
        return created, errors

    def _import_rooms(self, worksheet):
        created, errors = 0, []
        existing = {room.code for room in self.room_service.list_rooms()}
        for row_index, row in self._rows(worksheet):
            try:
                code = str(row.get("codigo") or "").strip().upper()
                if not code or code in existing:
                    continue
                self.room_service.create_room(
                    Room(None, code, str(row.get("nombre") or "").strip(), str(row.get("tipo") or "Ordinaria").strip(), int(row.get("capacidad") or 25), str(row.get("edificio") or "").strip(), str(row.get("recursos") or "").strip(), self._bool(row.get("disponible"), default=True))
                )
                existing.add(code)
                created += 1
            except Exception as exc:
                errors.append(f"Aulas fila {row_index}: {exc}")
        return created, errors

    def _import_course_subjects(self, worksheet):
        created, errors = 0, []
        courses = self._course_map()
        subjects = self._subject_map()
        teachers = self._teacher_map()

        for row_index, row in self._rows(worksheet):
            try:
                course_code = str(row.get("curso_codigo") or "").strip()
                subject_code = str(row.get("materia_codigo") or "").strip().upper()
                teacher_code = str(row.get("profesor_codigo") or "").strip()

                course = courses.get(course_code)
                subject = subjects.get(subject_code)
                teacher = teachers.get(teacher_code) if teacher_code else None

                if not course:
                    errors.append(f"MateriasCurso fila {row_index}: curso no encontrado: {course_code}")
                    continue
                if not subject:
                    errors.append(f"MateriasCurso fila {row_index}: materia no encontrada: {subject_code}")
                    continue

                self.course_subject_service.create_assignment(
                    CourseSubject(None, course.id, subject.id, int(row.get("sesiones_semanales") or subject.weekly_sessions), teacher.id if teacher else None, str(row.get("tipo_aula") or subject.room_type or "").strip(), str(row.get("notas") or "").strip())
                )
                created += 1
            except Exception as exc:
                errors.append(f"MateriasCurso fila {row_index}: {exc}")
        return created, errors

    def _import_availability(self, worksheet):
        updated, errors = 0, []
        teachers = self._teacher_map()
        for row_index, row in self._rows(worksheet):
            try:
                teacher_code = str(row.get("profesor_codigo") or "").strip()
                teacher = teachers.get(teacher_code)
                if not teacher:
                    errors.append(f"Disponibilidad fila {row_index}: profesor no encontrado: {teacher_code}")
                    continue
                self.availability_service.set_status(
                    teacher.id,
                    int(row.get("dia") or 1),
                    int(row.get("periodo") or 1),
                    self._normalise_status(str(row.get("estado") or "AVAILABLE").strip().upper()),
                )
                updated += 1
            except Exception as exc:
                errors.append(f"Disponibilidad fila {row_index}: {exc}")
        return updated, errors

    def _import_dynamic_rules(self, worksheet):
        created, errors = 0, []
        existing = {rule.code for rule in self.dynamic_rule_service.list_rules()}

        for row_index, row in self._rows(worksheet):
            try:
                code = str(row.get("codigo") or "").strip().upper()
                if not code or code in existing:
                    continue
                self.dynamic_rule_service.create_rule(
                    DynamicRule(
                        id=None,
                        code=code,
                        name=str(row.get("nombre") or "").strip(),
                        scope=str(row.get("ambito") or "Centro").strip(),
                        rule_type=str(row.get("tipo") or "Nota personalizada").strip(),
                        priority=str(row.get("prioridad") or "Preferente").strip(),
                        target=str(row.get("afecta_a") or "").strip(),
                        day=self._optional_int(row.get("dia")),
                        period=self._optional_int(row.get("periodo")),
                        value=str(row.get("valor") or "").strip(),
                        active=self._bool(row.get("activa"), default=True),
                        notes=str(row.get("notas") or "").strip(),
                    )
                )
                existing.add(code)
                created += 1
            except Exception as exc:
                errors.append(f"Reglas fila {row_index}: {exc}")
        return created, errors

    def _normalise_status(self, status: str) -> AvailabilityStatus:
        mapping = {
            "DISPONIBLE": AvailabilityStatus.AVAILABLE,
            "AVAILABLE": AvailabilityStatus.AVAILABLE,
            "PREFERENTE": AvailabilityStatus.PREFERRED,
            "PREFERRED": AvailabilityStatus.PREFERRED,
            "EVITAR": AvailabilityStatus.AVOID,
            "AVOID": AvailabilityStatus.AVOID,
            "NO DISPONIBLE": AvailabilityStatus.FORBIDDEN,
            "FORBIDDEN": AvailabilityStatus.FORBIDDEN,
        }
        return mapping.get(status, AvailabilityStatus.AVAILABLE)

    def _bool(self, value, default=False) -> bool:
        if value is None or str(value).strip() == "":
            return default
        return str(value).strip().casefold() in {"si", "sí", "yes", "true", "1"}

    def _optional_int(self, value):
        if value is None or str(value).strip() == "":
            return None
        return int(value)
